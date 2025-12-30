import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse # pyright: ignore[reportMissingImports]
from fastapi import UploadFile, File, HTTPException # pyright: ignore[reportMissingImports]
import numpy as np
from typing import List
from io import BytesIO

async def ValidationReceipt(
    transfers_files: List[UploadFile] = File(...),          
    ies_files: List[UploadFile] = File(...),        
):
    # Verificación básica de archivos
    if not transfers_files or not ies_files:
        raise HTTPException(status_code=400, detail="Debe proporcionar ambos tipos de archivos")

    try:
        # Procesar todos los archivos transfers
        all_transfers_dfs = []
        
        for transfers_file in transfers_files:
            transfers_content = await transfers_file.read()
            df_transfers = pd.read_excel(BytesIO(transfers_content))
            
            # Verificar que tenga las columnas necesarias
            required_columns = ['Inventory ID', 'Description', 'Quantity']
            if not all(col in df_transfers.columns for col in required_columns):
                raise HTTPException(
                    status_code=400, 
                    detail=f"Archivo {transfers_file.filename} no tiene las columnas requeridas: {required_columns}"
                )
            
            df_transfers = df_transfers[['Inventory ID', 'Description', 'Quantity']]
            df_transfers = df_transfers.groupby(['Inventory ID', 'Description'], as_index=False)['Quantity'].sum()
            all_transfers_dfs.append(df_transfers)
        
        # Combinar todos los DataFrames de transfers
        if all_transfers_dfs:
            combined_transfers = pd.concat(all_transfers_dfs, ignore_index=True)
            # Agrupar nuevamente por si hay duplicados entre archivos
            combined_transfers = combined_transfers.groupby(['Inventory ID', 'Description'], as_index=False)['Quantity'].sum()
        else:
            combined_transfers = pd.DataFrame(columns=['Inventory ID', 'Description', 'Quantity'])
        
        # Procesar todos los archivos IES
        all_ies_dfs = []
        
        for ies_file in ies_files:
            ies_content = await ies_file.read()
            df_ies = pd.read_excel(BytesIO(ies_content), skiprows=8)
            
            # Verificar que tenga las columnas necesarias
            required_columns_ies = ['NÚMERO DE PARTE', 'DESCRIPCIÓN EN INGLÉS', 'CANTIDAD']
            if not all(col in df_ies.columns for col in required_columns_ies):
                raise HTTPException(
                    status_code=400, 
                    detail=f"Archivo {ies_file.filename} no tiene las columnas requeridas: {required_columns_ies}"
                )
            
            df_ies = df_ies[['NÚMERO DE PARTE', 'DESCRIPCIÓN EN INGLÉS', 'CANTIDAD']]
            all_ies_dfs.append(df_ies)
        
        # Combinar todos los DataFrames de IES
        if all_ies_dfs:
            combined_ies = pd.concat(all_ies_dfs, ignore_index=True)
            # Agrupar por si hay duplicados entre archivos
            combined_ies = combined_ies.groupby(['NÚMERO DE PARTE', 'DESCRIPCIÓN EN INGLÉS'], as_index=False)['CANTIDAD'].sum()
        else:
            combined_ies = pd.DataFrame(columns=['NÚMERO DE PARTE', 'DESCRIPCIÓN EN INGLÉS', 'CANTIDAD'])
        combined_ies = combined_ies.rename(columns={
            'NÚMERO DE PARTE': 'Inventory ID',
            'DESCRIPCIÓN EN INGLÉS': 'Description',
            'CANTIDAD' : 'Quantity'
        })
        combined_ies = combined_ies.groupby(['Inventory ID','Description'], as_index=False)['Quantity'].sum()
        combined_ies['Inventory ID'] =  combined_ies['Inventory ID'].str.replace("'","",regex=False)
        #print(dfies)

        #Merge
        dfmerge = pd.merge(combined_transfers, combined_ies, on='Inventory ID', how='outer')
        dfmerge['Diference'] = dfmerge['Quantity_x'].fillna(0) - dfmerge['Quantity_y'].fillna(0)
        conditions = [
            dfmerge['Diference'].isna(),
            dfmerge['Diference'] > 0,
            dfmerge['Diference'] == 0,
            dfmerge['Diference'] < 0
        ]
        choices = ['Revisar', 'Revisar','Correcto', 'Revisar']
        dfmerge['status'] = np.select(conditions, choices)
        dfmerge = dfmerge.rename(columns={
            'Description_x':'Description',
            'Quantity_x':'Quantity Transfers',
            'Quantity_y': 'Quantity IES',
            'status': 'Comments'
        })
        dfmerge = dfmerge[['Inventory ID', 'Description', 'Quantity Transfers', 'Quantity IES', 'Diference','Comments']]

        # Create an Excel file with multiple sheet
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dfmerge.to_excel(writer, sheet_name='Data', index=False) 

            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # Definir colores
            COLORES = {
                'Correcto': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'Revisar': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            }
            
            # Encontrar columna 'Status'
            status_col_idx = None
            for idx, col in enumerate(dfmerge.columns, 1):
                if col == 'Comments':
                    status_col_idx = idx
                    break
            
            # Aplicar colores a la columna Status
            if status_col_idx:
                for row in range(2, len(dfmerge) + 2):
                    cell = worksheet.cell(row=row, column=status_col_idx)
                    status_value = cell.value
                    
                    if status_value in COLORES:
                        cell.fill = COLORES[status_value]
                
                # Centrar y poner en negrita
                for row in range(1, len(dfmerge) + 2):
                    cell = worksheet.cell(row=row, column=status_col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row > 1:  # Solo negrita para datos, no encabezados
                        cell.font = Font(bold=True)
            
            # Autoajustar columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename="Analisis_Importaciones.xlsx"'
            }
        )
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Error de codificación.")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Error al analizar el US Chase.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")