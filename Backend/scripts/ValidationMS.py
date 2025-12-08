import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse # pyright: ignore[reportMissingImports]
from fastapi import UploadFile, File, HTTPException # pyright: ignore[reportMissingImports]
import numpy as np

async def ValidationMS(
    file: UploadFile = File(...),          
    file2: UploadFile = File(...),        
):

    # Verificación básica de archivos
    if not file.filename or not file2.filename:
        raise HTTPException(status_code=400, detail="Debe proporcionar ambos archivos")

    try:
        # Leer contenido de los archivos
        tetakawifile = await file.read()
        manifest = await file2.read()

        #Manifest File
        nombre_hoja='Hoja1'
        dfmanifest = pd.read_excel(manifest, sheet_name=nombre_hoja, header=0)
        dfmanifestcol= dfmanifest[['No Parte', 'Cantidad']]
        dfmanifest_renombrado = dfmanifestcol.rename(columns={
            'No Parte': 'numero_parte',
            'Cantidad': 'cantidad'
        })


        #Tetakawi File
        dftetakawi = pd.read_excel(tetakawifile, skiprows=8)
        dftetakawicol = dftetakawi[['No. PARTE','TIPO IMPORTACIÓN','CANTIDAD ACTUAL']]
        dftkw_renombrado = dftetakawicol.rename(columns={
            'No. PARTE': 'numero_parte',
            'TIPO IMPORTACIÓN': 'tipo_importacion',
            'CANTIDAD ACTUAL':'cantidad'
        })
        dftkwgroup = dftkw_renombrado.groupby(['numero_parte','tipo_importacion'], as_index=False)['cantidad'].sum()
        dftkwgroup['numero_parte'] = dftkwgroup['numero_parte'].str.replace("'","",regex=False)

        dftemporal = dftkwgroup[dftkwgroup['tipo_importacion']=='Temporal / Temporary'].copy()
        dfdefinitivo = dftkwgroup[dftkwgroup['tipo_importacion']=='Definitiva / Definitive'].copy()
        dfdefinitivo_sin_ms = dftkwgroup[dftkwgroup['tipo_importacion']=='Definitiva / Definitive'].copy()
        dfdefinitivo_sin_ms['numero_parte'] = dfdefinitivo_sin_ms['numero_parte'] + '-MS'
        dfdefinitivo_sin_ms= dfdefinitivo_sin_ms[['numero_parte','cantidad']]
        dfdefinitivo_sin_ms = dfdefinitivo_sin_ms.rename(columns={
            'cantidad' :'cantidad_sinms'
        })
        dfmerge = pd.merge(dfmanifest_renombrado,dftemporal, on='numero_parte', how='left')
        dfmerge = pd.merge(dfmerge,dfdefinitivo,on='numero_parte', how='left')
        dfmerge = pd.merge(dfmerge,dfdefinitivo_sin_ms,on='numero_parte', how='left')
        dfmerge_renombrado = dfmerge.rename(columns={
            'cantidad_x': 'cantidad_solicitada',
            'cantidad_y': 'temporal',
            'cantidad': 'definitiva',
            'cantidad_sinms': 'definitiva_sin_ms'
        })
        dfmergecol = dfmerge_renombrado[['numero_parte', 'cantidad_solicitada', 'temporal', 'definitiva','definitiva_sin_ms']]
        dfmergecol = dfmergecol.fillna(0)
        dfmergecol['resta_temporal'] = dfmergecol['temporal'] - dfmergecol['cantidad_solicitada']
        # 1. Status inicial
        dfmergecol['status'] = np.where(dfmergecol['resta_temporal'] >= 0, 'Correcto', 'Revisar')

        def crear_mensaje_detallado(row):
            if row['status'] == 'Correcto':
                return ''
            
            deficit = -row['resta_temporal']
            
            # Calcular uso de recursos
            usado_def = min(row['definitiva'], deficit)
            deficit_restante = deficit - usado_def
            usado_sin_ms = min(row['definitiva_sin_ms'], deficit_restante)
            deficit_final = deficit_restante - usado_sin_ms
            
            # Calcular saldos para importar
            importar_def = row['definitiva'] - usado_def
            importar_sin_ms = row['definitiva_sin_ms'] - usado_sin_ms
            
            # Construir mensaje
            partes = []
            
            if importar_def > 0:
                partes.append(f"{int(importar_def)} como definitiva")
            if usado_def > 0:
                partes.append(f"({int(usado_def)} exportar como DEFINITIVA COM -MS)")
            
            if importar_sin_ms > 0:
                partes.append(f"{int(importar_sin_ms)} como definitiva sin ms")
            if usado_sin_ms > 0:
                partes.append(f"({int(usado_sin_ms)} exportar como DEFINITIVA SIN -MS)")
            
            if deficit_final > 0:
                partes.append(f"{int(deficit_final)} faltantes - SALDO INSUFICIENTE")
            
            return " ".join(partes)

        dfmergecol['mensaje_detallado'] = dfmergecol.apply(crear_mensaje_detallado, axis=1)

        # Actualizar status final
        dfmergecol['status_final'] = dfmergecol.apply(
            lambda row: 'Correcto' if row['status'] == 'Correcto' else 
                    ('Saldo Insuficiente' if row['mensaje_detallado'].endswith('SALDO INSUFICIENTE') 
                        else 'Exportación Especial'),
            axis=1
        )
        dfmergecol = dfmergecol[['numero_parte','cantidad_solicitada', 'temporal','definitiva','definitiva_sin_ms','status_final','mensaje_detallado']]
        dfmergecol_renombrado = dfmergecol.rename(columns={
            'numero_parte': 'No. Parte',
            'canidad_solicitada': 'Cantidad Solicitada',
            'temporal': 'Saldo Temporal',
            'definitiva': 'Saldo Definitivo',
            'definitiva_sin_ms': 'Saldo Def. sin -MS',
            'status_final': 'Status',
            'mensaje_detallado': 'Comentarios'
        })
        # Create an Excel file with multiple sheet
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dfmergecol_renombrado.to_excel(writer, sheet_name='Data', index=False)
            

            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            # Definir colores
            COLORES = {
                'Correcto': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'Exportación Especial': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'Saldo Insuficiente': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            }
            
            # Encontrar columna 'Status'
            status_col_idx = None
            for idx, col in enumerate(dfmergecol_renombrado.columns, 1):
                if col == 'Status':
                    status_col_idx = idx
                    break
            
            # Aplicar colores a la columna Status
            if status_col_idx:
                for row in range(2, len(dfmergecol_renombrado) + 2):
                    cell = worksheet.cell(row=row, column=status_col_idx)
                    status_value = cell.value
                    
                    if status_value in COLORES:
                        cell.fill = COLORES[status_value]
                
                # Centrar y poner en negrita
                for row in range(1, len(dfmergecol_renombrado) + 2):
                    cell = worksheet.cell(row=row, column=status_col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    if row > 1:  # Solo negrita para datos, no encabezados
                        cell.font = Font(bold=True)
            
            # Autoajustar columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': 'attachment; filename="Analisis_Importaciones.xlsx"'
            }
        )
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Error de codificación.")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Error al analizar el US Chase.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error inesperado: {str(e)}")